#!/usr/bin/env python3
"""Third-party XML benchmarks for the xml.etree serializer experiment."""

from __future__ import annotations

import argparse
import io
import json
import statistics
import sys
import time
from collections.abc import Callable

from openpyxl import Workbook
from openpyxl import __version__ as OPENPYXL_VERSION
from openpyxl import LXML
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def mixed_row(row_idx: int, cols: int) -> list[object]:
    row: list[object] = []
    for col_idx in range(1, cols + 1):
        if (row_idx + col_idx) % 10 == 0:
            row.append(f"text-{row_idx}-{col_idx}")
        else:
            row.append(row_idx * col_idx)
    return row


def save_workbook(workbook: Workbook) -> int:
    buffer = io.BytesIO()
    workbook.save(buffer)
    try:
        workbook.close()
    except AttributeError:
        pass
    return buffer.tell()


def build_basic_workbook(*, write_only: bool = False) -> Workbook:
    rows = 1000
    cols = 50
    sheets = 1
    workbook = Workbook(write_only=write_only)

    for sheet_idx in range(sheets):
        if write_only:
            sheet = workbook.create_sheet(title=f"Sheet{sheet_idx + 1}")
        else:
            if sheet_idx == 0:
                sheet = workbook.active
                sheet.title = "Sheet1"
            else:
                sheet = workbook.create_sheet(title=f"Sheet{sheet_idx + 1}")

        for row_idx in range(1, rows + 1):
            row = mixed_row(row_idx, cols)
            if write_only:
                sheet.append(row)
            else:
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

    return workbook


def bench_openpyxl_basic() -> int:
    return save_workbook(build_basic_workbook(write_only=False))


def bench_openpyxl_write_only() -> int:
    return save_workbook(build_basic_workbook(write_only=True))


def bench_openpyxl_styles() -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Styles"

    for row_idx in range(1, 251):
        for col_idx in range(1, 41):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
            color = f"{(row_idx * 7 + col_idx * 13) % 256:02X}"
            inverse = f"{255 - int(color, 16):02X}"
            cell.font = Font(
                name="Calibri",
                size=10 + (row_idx % 3),
                bold=(col_idx % 2 == 0),
                italic=(row_idx % 5 == 0),
                color=f"FF{color}{inverse}{color}",
            )
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=f"FF{color}{color}{color}",
            )
            cell.alignment = Alignment(
                horizontal="center" if col_idx % 3 == 0 else "left",
                vertical="center",
                wrap_text=(col_idx % 4 == 0),
            )
    return save_workbook(workbook)


def bench_openpyxl_comments() -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comments"

    for row_idx in range(1, 251):
        row = mixed_row(row_idx, 20)
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if (row_idx + col_idx) % 2 == 0:
                cell.comment = Comment(
                    text=(
                        f"Comment for row {row_idx}, column {col_idx}. "
                        f"Payload {row_idx * col_idx}."
                    ),
                    author=f"author-{col_idx % 7}",
                )

    return save_workbook(workbook)


def _populate_chart_sheet(sheet, *, rows: int, cols: int) -> None:
    headers = ["index"] + [f"series_{idx}" for idx in range(1, cols)]
    sheet.append(headers)
    for row_idx in range(1, rows + 1):
        row = [row_idx]
        for col_idx in range(1, cols):
            row.append(row_idx * col_idx + (row_idx % (col_idx + 3)))
        sheet.append(row)


def bench_openpyxl_charts() -> int:
    workbook = Workbook()
    sheets = [workbook.active] + [workbook.create_sheet() for _ in range(3)]

    for idx, sheet in enumerate(sheets, 1):
        sheet.title = f"Chart{idx}"
        _populate_chart_sheet(sheet, rows=400, cols=6)

        line = LineChart()
        line.title = f"Lines {idx}"
        line.y_axis.title = "Value"
        line.x_axis.title = "Index"
        data = Reference(sheet, min_col=2, max_col=6, min_row=1, max_row=401)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=401)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 7
        line.width = 15
        sheet.add_chart(line, "H2")

        bar = BarChart()
        bar.title = f"Bars {idx}"
        bar.type = "col"
        bar.style = 10
        data2 = Reference(sheet, min_col=2, max_col=4, min_row=1, max_row=201)
        cats2 = Reference(sheet, min_col=1, min_row=2, max_row=201)
        bar.add_data(data2, titles_from_data=True)
        bar.set_categories(cats2)
        sheet.add_chart(bar, "H20")

    return save_workbook(workbook)


def bench_openpyxl_tables() -> int:
    workbook = Workbook()
    sheets = [workbook.active] + [workbook.create_sheet() for _ in range(2)]

    for idx, sheet in enumerate(sheets, 1):
        sheet.title = f"Table{idx}"
        headers = [f"col_{col_idx}" for col_idx in range(1, 11)]
        sheet.append(headers)
        for row_idx in range(1, 801):
            sheet.append(mixed_row(row_idx, 10))

        table = Table(displayName=f"TableBench{idx}", ref="A1:J801")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    return save_workbook(workbook)


SCENARIOS: dict[str, Callable[[], int]] = {
    "openpyxl-basic": bench_openpyxl_basic,
    "openpyxl-write-only": bench_openpyxl_write_only,
    "openpyxl-styles": bench_openpyxl_styles,
    "openpyxl-comments": bench_openpyxl_comments,
    "openpyxl-charts": bench_openpyxl_charts,
    "openpyxl-tables": bench_openpyxl_tables,
}


def run_benchmark(func: Callable[[], int], *, repeat: int, warmup: int) -> dict[str, object]:
    for _ in range(warmup):
        func()

    runs_ms: list[float] = []
    result_size = 0
    for _ in range(repeat):
        start = time.perf_counter()
        result_size = func()
        elapsed = time.perf_counter() - start
        runs_ms.append(elapsed * 1000.0)

    stats: dict[str, object] = {
        "runs_ms": runs_ms,
        "mean_ms": statistics.mean(runs_ms),
        "median_ms": statistics.median(runs_ms),
        "min_ms": min(runs_ms),
        "max_ms": max(runs_ms),
        "result_bytes": result_size,
    }
    if len(runs_ms) > 1:
        stats["stdev_ms"] = statistics.stdev(runs_ms)
    else:
        stats["stdev_ms"] = 0.0
    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--scenario",
        choices=["all", *SCENARIOS.keys()],
        default="all",
        help="benchmark scenario to run",
    )
    parser.add_argument("--repeat", type=int, default=5, help="measured repetitions")
    parser.add_argument("--warmup", type=int, default=1, help="warmup repetitions")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    selected = (
        list(SCENARIOS.items())
        if args.scenario == "all"
        else [(args.scenario, SCENARIOS[args.scenario])]
    )

    payload = {
        "python": sys.version,
        "openpyxl_version": OPENPYXL_VERSION,
        "openpyxl_lxml": LXML,
        "repeat": args.repeat,
        "warmup": args.warmup,
        "benchmarks": {},
    }

    for name, func in selected:
        payload["benchmarks"][name] = run_benchmark(
            func,
            repeat=args.repeat,
            warmup=args.warmup,
        )

    json.dump(payload, sys.stdout, indent=2, sort_keys=True)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
